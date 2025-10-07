"""Utilities for rendering single-file XLSX reports."""

from __future__ import annotations

from collections.abc import Iterable
from os import PathLike
from typing import TYPE_CHECKING, Any, BinaryIO

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

if TYPE_CHECKING:  # pragma: no cover - import for typing only
    from .diff import DiffResult


def _append_rows(ws, rows: Iterable[tuple[Any, ...]]) -> None:
    for row in rows:
        ws.append(list(row))


def _set_table_formatting(ws) -> None:  # type: ignore[no-untyped-def]
    ws.freeze_panes = "A2"
    last_column = get_column_letter(ws.max_column or 1)
    last_row = ws.max_row or 1
    ws.auto_filter.ref = f"A1:{last_column}{last_row}"


def _format_summary_sheet(ws, payload: dict[str, Any]) -> None:  # type: ignore[no-untyped-def]
    meta = payload.get("meta", {})
    summary = payload.get("summary", {})
    thresholds = summary.get("thresholds", {})

    rows = [
        ("Metric", "Value"),
        ("Table", meta.get("table", "")),
        ("Schema", meta.get("schema", "")),
        ("Before file", meta.get("before_file", "")),
        ("After file", meta.get("after_file", "")),
        ("Jira", meta.get("jira", "")),
        ("Added", summary.get("added", 0)),
        ("Removed", summary.get("removed", 0)),
        ("Changed", summary.get("changed", 0)),
        ("Total before", summary.get("total_before", 0)),
        ("Total after", summary.get("total_after", 0)),
        (
            "Unexpected partners",
            ", ".join(summary.get("unexpected_partners", [])) or "None",
        ),
        ("Max added", thresholds.get("max_added", "")),
        ("Max removed", thresholds.get("max_removed", "")),
        (
            "Threshold violations",
            ", ".join(thresholds.get("violations", [])) or "None",
        ),
    ]

    _append_rows(ws, rows)
    _set_table_formatting(ws)


def _collect_headers(records: list[dict[str, Any]], key: str) -> list[str]:
    headers: list[str] = ["key"]
    for record in records:
        section = record.get(key, {})
        if not isinstance(section, dict):
            continue
        for field_name in section:
            if field_name not in headers:
                headers.append(field_name)
    return headers


def _write_record_sheet(ws, records: list[dict[str, Any]], section_key: str) -> None:  # type: ignore[no-untyped-def]
    headers = _collect_headers(records, section_key)
    ws.append(headers)
    for record in records:
        row = [record.get("key", "")]
        section = record.get(section_key, {})
        values = section if isinstance(section, dict) else {}
        for header in headers[1:]:
            row.append(values.get(header, ""))
        ws.append(row)
    _set_table_formatting(ws)


def _write_changed_sheet(ws, records: list[dict[str, Any]]) -> None:  # type: ignore[no-untyped-def]
    ws.append(["key", "field", "before", "after"])
    for record in records:
        changes = record.get("changes", {})
        if not isinstance(changes, dict):
            continue
        for field_name, change in sorted(changes.items()):
            before_value = ""
            after_value = ""
            if isinstance(change, dict):
                before_value = change.get("before", "")
                after_value = change.get("after", "")
            ws.append([record.get("key", ""), field_name, before_value, after_value])
    _set_table_formatting(ws)


def render_xlsx_report(result: DiffResult, payload: dict[str, Any]) -> Workbook:
    """Create an openpyxl workbook representing the diff result."""

    workbook = Workbook()
    summary_ws = workbook.active
    summary_ws.title = "Summary"
    _format_summary_sheet(summary_ws, payload)

    added_ws = workbook.create_sheet("Added")
    _write_record_sheet(added_ws, result.added, "after")

    removed_ws = workbook.create_sheet("Removed")
    _write_record_sheet(removed_ws, result.removed, "before")

    changed_ws = workbook.create_sheet("Changed")
    _write_changed_sheet(changed_ws, result.updated)

    return workbook


def write_xlsx_report(
    result: DiffResult,
    payload: dict[str, Any],
    destination: BinaryIO | PathLike[str] | str,
) -> None:
    """Render the XLSX workbook and write it to a file-like object or path."""

    workbook = render_xlsx_report(result, payload)
    workbook.save(destination)
