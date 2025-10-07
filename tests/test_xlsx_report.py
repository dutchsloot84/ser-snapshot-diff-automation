from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from serdiff import cli
from serdiff.diff import build_canonical_payload, diff_files, write_reports
from serdiff.presets import get_preset
from serdiff.report_xlsx import write_xlsx_report


def write_xml(path: Path, body: str) -> Path:
    path.write_text(body.strip(), encoding="utf-8")
    return path


def _lookup_summary_value(sheet, label: str) -> str:  # type: ignore[no-untyped-def]
    for row in sheet.iter_rows(min_row=2, values_only=True):
        metric = row[0]
        value = row[1] if len(row) > 1 else None
        if metric == label:
            return value
    raise AssertionError(f"Label {label!r} not found in summary sheet")


def test_xlsx_report_generates_expected_workbook(tmp_path: Path) -> None:
    before = write_xml(
        tmp_path / "before.xml",
        """
        <ExposureTypes>
          <ExposureType>
            <PublicID>EXP-001</PublicID>
            <ExposureCode>AUTO</ExposureCode>
            <Partner>Uber</Partner>
            <State>NY</State>
          </ExposureType>
        </ExposureTypes>
        """,
    )
    after = write_xml(
        tmp_path / "after.xml",
        """
        <ExposureTypes>
          <ExposureType>
            <PublicID>EXP-001</PublicID>
            <ExposureCode>AUTO</ExposureCode>
            <Partner>Lyft</Partner>
            <State>CA</State>
          </ExposureType>
          <ExposureType>
            <PublicID>EXP-002</PublicID>
            <ExposureCode>AUTO</ExposureCode>
            <Partner>DoorDash</Partner>
            <State>TX</State>
          </ExposureType>
        </ExposureTypes>
        """,
    )

    preset = get_preset("EXPOSURE")
    result = diff_files(
        before,
        after,
        preset.config,
        jira="MOB-555",
        expected_partners=["Uber"],
    )

    thresholds, _ = cli._evaluate_thresholds(
        result,
        expected_partners=["Uber"],
        max_added=0,
        max_removed=0,
    )

    payload = build_canonical_payload(result, thresholds)

    buffer = BytesIO()
    write_xlsx_report(result, payload, buffer)
    buffer.seek(0)
    workbook = load_workbook(filename=buffer, data_only=True)

    assert workbook.sheetnames == ["Summary", "Added", "Removed", "Changed"]

    summary = workbook["Summary"]
    assert summary.freeze_panes == "A2"
    assert summary.auto_filter.ref == f"A1:B{summary.max_row}"
    assert _lookup_summary_value(summary, "Added") == payload["summary"]["added"]
    assert _lookup_summary_value(summary, "Removed") == payload["summary"]["removed"]
    assert _lookup_summary_value(summary, "Changed") == payload["summary"]["changed"]
    assert "DoorDash" in _lookup_summary_value(summary, "Unexpected partners")
    assert "UNEXPECTED_PARTNER" in _lookup_summary_value(summary, "Threshold violations")

    added = workbook["Added"]
    assert added.freeze_panes == "A2"
    assert added.auto_filter.ref.startswith("A1:")
    header = next(added.iter_rows(min_row=1, max_row=1, values_only=True))
    expected_headers = ("key", *tuple(preset.config.fields))
    assert header == expected_headers
    added_row = next(added.iter_rows(min_row=2, max_row=2, values_only=True))
    assert added_row[0].startswith("PublicID=EXP-002")
    partner_column = header.index("Partner")
    assert added_row[partner_column] == "DoorDash"

    changed = workbook["Changed"]
    assert changed.freeze_panes == "A2"
    assert changed.auto_filter.ref.startswith("A1:")
    changed_rows = list(changed.iter_rows(min_row=2, values_only=True))
    assert ("PublicID=EXP-001", "Partner", "Uber", "Lyft") in changed_rows
    assert ("PublicID=EXP-001", "State", "NY", "CA") in changed_rows

    out_dir = tmp_path / "reports" / "run"
    produced = write_reports(
        result,
        out_dir,
        report_type="xlsx",
        thresholds=thresholds,
    )

    json_path = out_dir / "diff.json"
    xlsx_path = out_dir / f"{out_dir.name}.xlsx"

    assert str(json_path) in produced
    assert str(xlsx_path) in produced
    assert all(not path.endswith(".csv") for path in produced)
